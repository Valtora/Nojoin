"""Spreadsheet and CSV parsing.

One page per worksheet, which makes the sheet name a real page title and keeps
an unrelated tab from diluting the retrieval unit of the one being asked about.
Rows are not truncated: a spreadsheet attached to a meeting is usually attached
precisely because someone wants a figure out of it, and a silently cut sheet is
worse than a large one. Oversized sheets are split by the chunker instead.
"""

from __future__ import annotations

import csv
import logging
from typing import Iterator, List, Optional

from .markdown_tables import rows_to_markdown_table
from .types import DocumentSource, PageSource

logger = logging.getLogger(__name__)

# csv.field_size_limit defaults low enough that one long cell aborts the parse.
# Raised rather than removed, since a genuinely unbounded field is a sign of a
# malformed file rather than of legitimate data.
CSV_FIELD_SIZE_LIMIT = 10 * 1024 * 1024


def open_xlsx(path: str, *, want_images: bool) -> DocumentSource:
    from openpyxl import load_workbook

    # read_only streams rows instead of building a cell object graph, and
    # data_only takes cached formula results -- a formula string is noise to a
    # language model, its computed value is the answer.
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_names = list(workbook.sheetnames)

    def pages() -> Iterator[PageSource]:
        try:
            for index, name in enumerate(sheet_names):
                worksheet = workbook[name]
                rows: List[List[object]] = []
                for row in worksheet.iter_rows(values_only=True):
                    if row is None:
                        continue
                    if all(value is None or str(value).strip() == "" for value in row):
                        continue
                    rows.append(list(row))
                yield PageSource(
                    page_number=index + 1,
                    title=name,
                    text=rows_to_markdown_table(rows),
                )
        finally:
            workbook.close()

    return DocumentSource(page_count=len(sheet_names), pages=pages)


def _sniff_dialect(sample: str) -> Optional[type]:
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        # Sniffing fails on single-column files, among others. The default
        # comma dialect reads those correctly, so this is not worth an error.
        return None


def open_csv(path: str, *, want_images: bool) -> DocumentSource:
    csv.field_size_limit(CSV_FIELD_SIZE_LIMIT)

    with open(path, "r", encoding="utf-8", errors="replace", newline="") as handle:
        sample = handle.read(64 * 1024)
        handle.seek(0)
        dialect = _sniff_dialect(sample)
        reader = csv.reader(handle, dialect) if dialect else csv.reader(handle)
        rows = [row for row in reader if any(cell.strip() for cell in row)]

    table = rows_to_markdown_table(rows)

    def pages() -> Iterator[PageSource]:
        yield PageSource(page_number=1, title=None, text=table)

    return DocumentSource(page_count=1, pages=pages)
